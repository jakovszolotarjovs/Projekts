import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By

import time

import random
import string

from openpyxl import Workbook, load_workbook


service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service = service, options = option)


myLogin = "jakovs-zolotarjovs-hotel@inbox.lv" #E-pasta logins
myPassword = "Hotel-Jakovs12345" #E-pasta parole



wb = load_workbook('clientData.xlsx') #Darbs ar XLSX failu
ws = wb.active
max_row = ws.max_row
sheet = wb.active
nameList = []
emailList = []
codeList = []
for i in range(2, max_row + 1):
    name = (ws['A'+ str(i)].value)
    email = (ws['B'+ str(i)].value)
    #ģenerē randomu piedavājuma kodu
    disCode = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
    #pieraksta piedavājuma kodu exel dokumentā
    tableDis = sheet.cell(row=i, column=3)
    tableDis.value=disCode
    if type(name)==str  and type(email)==str and type(disCode)==str:
        nameList.append(name)
        emailList.append(email)
        codeList.append(disCode)
wb.save("clientData.xlsx")


#Darbs ar e-pastu izmantojot selenium biblioteku
url = "https://www.inbox.lv/" 
driver.get(url) #Atveram E-pastu
time.sleep(12)#gaidam kamēr pabeigsies reklāma
find = driver.find_element(By.ID, "imapuser")
find.send_keys(myLogin)
find = driver.find_element(By.ID, "pass")
find.send_keys(myPassword)
find = driver.find_element(By.ID, "btn_sign-in")
find.click()
time.sleep(3)#gaidam kamēr ieladēsies lapa

for i in range(0, len(nameList)):#Rakstam personalizēto vēstuli
    time.sleep(0.3)
    find=driver.find_element(By.ID, "mail-menu__link_compose")
    find.click()
    find=driver.find_element(By.ID, "suggest-to")
    find.send_keys(emailList[i])

    find=driver.find_element(By.ID, "subject")
    find.send_keys(nameList[i] + ", piedavajums tieši jums!")

    #Meklējam iframe bazejoties uz klases nosaukumu
    iframe = driver.find_element(By.CLASS_NAME, "cke_wysiwyg_frame")
    #Pārsledzamies uz iframe
    driver.switch_to.frame(iframe)
    find = driver.find_element(By.TAG_NAME, "body")
    find.click()
    find.send_keys("Labdien,"+ nameList[i]+"! " + "\n\n" + 
                    " Specialais piedavajums tieši Jums. Tikai šajā menesī, īrēiet istabiņu " +
                    " musū viesnīcā uz 2 naktīm, un saņemiet masāžu un atsvaidzinošu tropisko " +
                    "kokteili bez maksas! Akcija spēkā vēl 2 nedēļas. Veicot "
                    "rezervāciju, neaizmirstiet pateikt savu unikālo kodu un uzrādīt "
                    " dokumentu."+ "\n\n"+ " Jūsu kupona kods ir "
                   + codeList[i]+"\n"+" Gaidīsim Jūs!)" + "\n\n"+ 
                   "Jā Jums rodas kaut-kādi jautājumi, rakstiet uz jakovs-zolotarjovs-hotel@inbox.lv")
    #aizveram iframe
    driver.switch_to.default_content()
    find = driver.find_element(By.CLASS_NAME, "Button-sc-3zooqt-0.chWxlP")#sūtam vēstuli
    find.click()



